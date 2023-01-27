from openpyxl import Workbook,load_workbook

wb= load_workbook("통합 문서1.xlsx")
ws=wb["Sheet1"]
body=open('outline.txt','r',encoding='utf-8')
message=body.read()
for x in range(5,77):
    email=ws[f"K{x}"].value
    firm=ws[f"H{x}"].value
    header=f"Dear {firm},\n\n"
    new_message=header+message
    ws[f"O{x}"]=new_message

wb.save('통합문서파이널.xlsx')
wb.close()
