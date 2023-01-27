from email.message import EmailMessage
from smtplib import SMTP_SSL
from openpyxl import Workbook,load_workbook

with SMTP_SSL("smtp.gmail.com",465) as smtp:
    smtp.login('kb081434@khu.ac.kr','matilda2021818')
    
    wb= load_workbook("통합문서파이널.xlsx")
    ws=wb["Sheet1"]
    for x in range(8,77):#5,77
        email=ws[f"K{x}"].value
        message=ws[f"O{x}"].value
        msg=EmailMessage()
        msg["To"]=email
        msg["Subject"]="[We invite you to PULSE9 in upcoming CES2023!]"
        msg.set_content(message)
        smtp.send_message(msg)
    wb.close()
