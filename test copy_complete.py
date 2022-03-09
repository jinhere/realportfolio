from openpyxl import Workbook,load_workbook

wb= load_workbook("3EA1.xlsx")# 시험지 이름 바꾸기
num=len(wb.sheetnames)
for x in range(num):
    wsname=wb.sheetnames[x]
    ws=wb[wsname]
    ws["C8"]=2021.12
    ws["D11"]=27
    ws["D12"]=26
 #   ws["D13"]=33
	
wb.save('3EA1.xlsx') # 이름 같게
wb.close()
