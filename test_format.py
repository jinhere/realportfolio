from openpyxl import Workbook,load_workbook
wb= load_workbook("절대어휘2.xlsx")# 시험지 이름 바꾸기
for x in range(1,31):
    ws=wb["day{}".format(x)]
    for y in range(5,34,2):
        ws["G{}".format(y)]=" "
        ws["O{}".format(y)]=" "
    for y in range(4,33,2):
        ws["H{}".format(y)]=" "
        ws["P{}".format(y)]=" "    
wb.save('modified1.xlsx')
wb.close()
